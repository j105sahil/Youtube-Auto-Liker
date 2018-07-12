import openpyxl
import webbrowser,sys
import pyautogui
from time import sleep
import random
import subprocess
chrome_path = 'C:/Program Files (x86)/Google/Chrome/Application/chrome.exe %s'
workbook = openpyxl.load_workbook('trial.xlsx')
sheet = workbook.get_sheet_by_name('Sheet1')
i=1
while True:
    if i%10==0:
        sleep(10)
        subprocess.call("taskkill /IM chrome.exe")

    if sheet.cell(row = i , column = 1).value!=None:
        a=random.randint(5,500)
        b=random.randint(5,500)
        webbrowser.get(chrome_path).open(sheet.cell(row = i , column = 1).value)
        pyautogui.moveTo(a,b,duration=3)
        #postion = pyautogui.locateCenterOnScreen('Capture.PNG')
        #pyautogui.click(postion)
    else:
        break
    sleep(5)
    i +=1
subprocess.call("taskkill /IM chrome.exe")
